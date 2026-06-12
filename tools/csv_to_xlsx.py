"""
将 CSV 文件转换为 XLSX 文件。

运行方式：
  python tools/csv_to_xlsx.py

启动后按提示输入 CSV 文件路径，支持：
  - 逐个文件输入，每个文件可自定义输出名称
  - 回车直接使用默认名称（同名 .xlsx）
  - 转换后询问是否继续，可反复转换
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def auto_column_width(worksheet, min_width: int = 8, max_width: int = 40) -> None:
    """根据内容自动调整列宽。"""
    for col_cells in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # 中文字符按 2 个字符宽度计算
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        adjusted = max(min_width, min(max_len + 2, max_width))
        worksheet.column_dimensions[col_letter].width = adjusted


def csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> Path:
    """将单个 CSV 转为 XLSX，返回输出路径。"""
    # 读取 CSV（兼容 utf-8-sig BOM 头）
    df = pd.read_csv(csv_path, encoding="utf-8-sig")

    # 写出 XLSX
    df.to_excel(xlsx_path, index=False, engine="openpyxl")

    # 自动调整列宽
    wb = load_workbook(xlsx_path)
    ws = wb.active
    auto_column_width(ws)
    wb.save(xlsx_path)

    print(f"  => {xlsx_path}  ({len(df)} 行 × {len(df.columns)} 列)")
    return xlsx_path


def resolve_output_name(csv_path: Path) -> Path | None:
    """
    询问用户输出文件名。
    - 回车：使用默认名称（同名 .xlsx）
    - 输入名称：可使用相对/绝对路径，不含扩展名则自动补 .xlsx
    - 输入 q：取消当前文件
    - 目标已存在时提示确认
    """
    default = csv_path.with_suffix(".xlsx")
    print(f"\n  源文件: {csv_path}")
    prompt = f"  输出文件名（回车默认 {default.name}，输入 q 跳过）: "
    raw = input(prompt).strip().strip('"').strip("'")

    if raw.lower() == "q":
        print("  [跳过] 用户取消")
        return None

    if not raw:
        xlsx_path = default
    else:
        xlsx_path = Path(raw)
        # 确保在脚本工作目录下
        if not xlsx_path.is_absolute():
            xlsx_path = Path.cwd() / xlsx_path
        # 自动补扩展名
        if xlsx_path.suffix.lower() != ".xlsx":
            xlsx_path = xlsx_path.with_suffix(".xlsx")

    # 目标已存在：提示覆盖确认
    if xlsx_path.exists():
        confirm = input(f"  [警告] 文件已存在: {xlsx_path.name}\n         是否覆盖？(y/n): ").strip().lower()
        if confirm != "y":
            print("  [跳过] 取消覆盖")
            return None

    return xlsx_path


def process_csv(csv_path: Path, xlsx_path: Path) -> bool:
    """处理单个 CSV 文件，成功返回 True。"""
    try:
        csv_to_xlsx(csv_path, xlsx_path)
        return True
    except Exception as e:
        print(f"  [错误] {csv_path.name}: {e}")
        return False


def prompt_for_file() -> Path | None:
    """询问用户输入单个 CSV 文件路径。"""
    while True:
        raw = input("\n请输入 CSV 文件路径（输入 q 退出）: ").strip().strip('"').strip("'")
        if raw.lower() == "q":
            return None
        if not raw:
            print("  [提示] 输入为空，请重新输入。")
            continue

        p = Path(raw)
        if not p.is_file():
            print(f"  [警告] 文件不存在: {p}")
            continue
        if p.suffix.lower() != ".csv":
            print(f"  [警告] 非 CSV 文件: {p}")
            continue

        return p


def main():
    print("=" * 50)
    print("  CSV → XLSX 转换工具")
    print("=" * 50)

    while True:
        csv_path = prompt_for_file()
        if csv_path is None:
            print("已退出。")
            break

        xlsx_path = resolve_output_name(csv_path)
        if xlsx_path is None:
            continue

        process_csv(csv_path, xlsx_path)

        again = input("\n是否继续转换其他文件？(y/n): ").strip().lower()
        if again != "y":
            print("已退出。")
            break


if __name__ == "__main__":
    main()
