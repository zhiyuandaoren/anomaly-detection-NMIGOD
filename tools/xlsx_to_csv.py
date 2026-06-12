#!/usr/bin/env python3
"""
Interactive script to convert XLSX files to CSV format.
User will be asked to choose mode and provide file/folder paths.
"""

import os
import sys
import csv

def convert_with_pandas(input_path, output_path):
    """Convert using pandas (fast and handles sheets)."""
    import pandas as pd
    df = pd.read_excel(input_path, sheet_name=0, engine='openpyxl')
    df.to_csv(output_path, index=False, encoding='utf-8')
    print(f"[OK] Converted: {input_path} -> {output_path}")
    return True

def convert_with_standard_lib(input_path, output_path):
    """Convert using openpyxl and csv (no pandas)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.stderr.write("Error: 'openpyxl' is required. Install with: pip install openpyxl\n")
        return False

    try:
        wb = load_workbook(input_path, read_only=True)
        ws = wb.active  # first sheet only

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        wb.close()
        print(f"[OK] Converted: {input_path} -> {output_path}")
        return True
    except Exception as e:
        sys.stderr.write(f"Error converting {input_path}: {e}\n")
        return False

def convert_file(input_path, output_path=None):
    """Convert a single xlsx file to csv."""
    if not os.path.exists(input_path):
        print(f"[FAIL] File not found: {input_path}")
        return False

    if output_path is None:
        base = os.path.splitext(input_path)[0]
        output_path = base + ".csv"

    # Try pandas first, fall back to openpyxl on any failure
    try:
        import pandas as pd
        return convert_with_pandas(input_path, output_path)
    except (ImportError, Exception) as e:
        if isinstance(e, ImportError):
            print("pandas not available, falling back to openpyxl...")
        else:
            print(f"pandas conversion failed ({e}), falling back to openpyxl...")
        return convert_with_standard_lib(input_path, output_path)

def convert_folder(folder_path):
    """Convert all .xlsx files in a folder to .csv files."""
    if not os.path.isdir(folder_path):
        print(f"[FAIL] Not a directory: {folder_path}")
        return

    xlsx_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xlsx')]
    if not xlsx_files:
        print(f"No .xlsx files found in {folder_path}")
        return

    print(f"Found {len(xlsx_files)} Excel file(s) in '{folder_path}':")
    for filename in xlsx_files:
        input_path = os.path.join(folder_path, filename)
        output_path = os.path.splitext(input_path)[0] + ".csv"
        convert_file(input_path, output_path)

def get_user_choice():
    """Display menu and get user choice."""
    print("\n" + "="*50)
    print("   XLSX to CSV Converter - Interactive Mode")
    print("="*50)
    print("1. Convert a single XLSX file")
    print("2. Convert all XLSX files in a folder")
    print("0. Exit")
    print("-"*50)

    while True:
        choice = input("Please enter your choice (0/1/2): ").strip()
        if choice in ('0', '1', '2'):
            return choice
        print("Invalid input. Please enter 0, 1 or 2.")

def main():
    while True:
        choice = get_user_choice()

        if choice == '0':
            print("Goodbye!")
            break

        elif choice == '1':
            # Single file mode
            file_path = input("Enter the path to the XLSX file: ").strip().strip('"')
            if not file_path:
                print("No path provided. Returning to menu.\n")
                continue

            # Ask for optional output path
            output_path = input("Enter output CSV path (press Enter to auto-generate): ").strip().strip('"')
            if not output_path:
                output_path = None

            print("\nConverting...")
            convert_file(file_path, output_path)
            print("\n" + "-"*50 + "\n")

        elif choice == '2':
            # Folder mode
            folder_path = input("Enter the folder path containing XLSX files: ").strip().strip('"')
            if not folder_path:
                print("No path provided. Returning to menu.\n")
                continue

            print("\nConverting...")
            convert_folder(folder_path)
            print("\n" + "-"*50 + "\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nInterrupted by user. Exiting.")
        sys.exit(0)